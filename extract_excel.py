import openpyxl
import os

file_path = '/Users/tomaszuscinski/Documents/Visual Code Studio/Kalkulator/Model oszczędności vs. WACC.xlsx'

if not os.path.exists(file_path):
    print(f"Error: File not found at {file_path}")
    exit(1)

try:
    wb = openpyxl.load_workbook(file_path, data_only=False)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n=== SHEET: {sheet_name} ===")
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    print(f"  {cell.coordinate}: {repr(cell.value)}")
except Exception as e:
    print(f"An error occurred: {e}")
